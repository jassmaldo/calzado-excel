import streamlit as st
import tempfile, os, re, io
import parsers

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    pass

# ── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Generador Excel · Calzado",
    page_icon="👟",
    layout="centered",
)

# ── Estilos personalizados ───────────────────────────────────────────────────
st.markdown("""
<style>
    .block-container { padding-top: 2rem; max-width: 780px; }
    .metric-card {
        background: #ffffff;
        border: 1px solid #E5E3DC;
        border-radius: 8px;
        padding: 12px 16px;
        text-align: center;
    }
    .metric-label { color: #8A8880; font-size: 12px; margin-bottom: 2px; }
    .metric-value { color: #1A1A1A; font-size: 22px; font-weight: 700; }
    .stAlert { border-radius: 8px; }
    div[data-testid="stFileUploader"] {
        border: 1.5px dashed #1D9E75;
        border-radius: 8px;
        padding: 8px;
    }
</style>
""", unsafe_allow_html=True)

# ── Cabecera ─────────────────────────────────────────────────────────────────
st.markdown("## 👟 Generador Excel · Calzado")
st.markdown("Sube los dos documentos del proveedor y descarga el Excel listo para importar.")
st.divider()

# ── Subida de archivos ───────────────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    st.markdown("**① Proforma / Invoice**")
    st.caption("Excel, PDF o Word · precios y referencias")
    inv_file = st.file_uploader("Invoice", type=["xlsx","xlsm","pdf","docx"],
                                 label_visibility="collapsed", key="inv")
with col2:
    st.markdown("**② Packing List**")
    st.caption("Excel, PDF o Word · cajas y distribución")
    pl_file  = st.file_uploader("Packing", type=["xlsx","xlsm","pdf","docx"],
                                 label_visibility="collapsed", key="pl")

# ── Lógica principal ─────────────────────────────────────────────────────────
if inv_file and pl_file:

    # Guardar en archivos temporales (parsers necesita rutas de disco)
    suffix_inv = os.path.splitext(inv_file.name)[1]
    suffix_pl  = os.path.splitext(pl_file.name)[1]

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix_inv) as ti:
        ti.write(inv_file.read()); inv_path = ti.name
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix_pl) as tp:
        tp.write(pl_file.read()); pl_path = tp.name

    # ── Parseo ───────────────────────────────────────────────────────────────
    with st.spinner("Analizando archivos..."):
        try:
            items, meta = parsers.parse(inv_path, pl_path)
        except Exception as e:
            st.error(f"Error al leer los archivos: {e}")
            st.stop()

    os.unlink(inv_path); os.unlink(pl_path)

    marca          = meta.get("marca") or meta.get("marca_archivo") or "SIN MARCA"
    uniform_tallas = meta.get("tallas")
    uniform_dist   = meta.get("dist")
    alertas        = meta.get("alertas", [])
    por_linea_marca= meta.get("por_linea_marca", False)

    total_pares = sum(i["pares"] for i in items if isinstance(i.get("pares"), int))
    total_cajas = sum(i["cajas"] for i in items if isinstance(i.get("cajas"), int))

    if uniform_tallas:
        tallas_txt = " · ".join(map(str, uniform_tallas))
        dist_txt   = " · ".join(map(str, uniform_dist))
    else:
        tallas_txt = "Variable por línea"
        dist_txt   = "Variable por línea"

    if por_linea_marca:
        marcas = list(dict.fromkeys(i.get("marca_linea","") for i in items if i.get("marca_linea")))
        marca_txt = " / ".join(marcas) if marcas else marca
    else:
        marca_txt = marca

    # ── Panel de resultados ──────────────────────────────────────────────────
    st.divider()
    st.success("✓ Datos detectados automáticamente")

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    def _metric(col, label, value):
        col.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
        </div>""", unsafe_allow_html=True)

    _metric(c1, "Marca",       marca_txt)
    _metric(c2, "Tallas",      tallas_txt)
    _metric(c3, "Distribución",dist_txt)
    _metric(c4, "Total pares", f"{total_pares:,}")
    _metric(c5, "Total cajas", f"{total_cajas:,}")
    _metric(c6, "Líneas",      str(len(items)))

    # ── Alertas ──────────────────────────────────────────────────────────────
    if alertas:
        st.divider()
        for a in alertas:
            if a.get("tipo") == "manual":
                st.warning(f"⚠️ {a['mensaje']}")
            else:
                st.info(f"ℹ️ {a['mensaje']}")

    # ── Preguntas opcionales ─────────────────────────────────────────────────
    has_color_q = any(re.match(r"^\d+-[A-Z]", i.get("color","")) for i in items)
    has_model_q = any("-" in i.get("ref","") and not re.match(r"^\d+-[A-Z]", i.get("ref",""))
                      for i in items)

    color_split = False
    model_split = False

    if has_color_q or has_model_q:
        st.divider()

    if has_color_q:
        ex = next(i["color"] for i in items if re.match(r"^\d+-[A-Z]", i.get("color","")))
        with st.expander("❓ Colores con guión detectados", expanded=True):
            st.caption(f"Ej: `{ex[:40]}`")
            color_split = st.radio(
                "¿Separar el número como Cód. Color y el texto como Color?",
                ["Sí, separar", "No, dejar igual"], key="cq", horizontal=True
            ) == "Sí, separar"

    if has_model_q:
        ex = next(i["ref"] for i in items if "-" in i.get("ref","")
                  and not re.match(r"^\d+-[A-Z]", i.get("ref","")))
        with st.expander("❓ Modelos con guión detectados", expanded=True):
            st.caption(f"Ej: `{ex}`")
            model_split = st.radio(
                "¿Separar el número final como Cód. Color y el resto como Modelo?",
                ["Sí, separar", "No, dejar igual"], key="mq", horizontal=True
            ) == "Sí, separar"

    # ── Vista previa ─────────────────────────────────────────────────────────
    st.divider()
    with st.expander("👁 Vista previa (primeras 10 líneas)", expanded=False):
        preview = []
        for item in items[:10]:
            if "modelo_forzado" in item:
                modelo    = item["modelo_forzado"]
                cod_color = item.get("cod_color_forzado", "")
                color     = item.get("color", "")
                m         = item.get("marca_linea", marca)
            else:
                modelo, color, cod_color = item["ref"], item.get("color",""), ""
                m = marca
                if model_split and "-" in modelo:
                    parts = modelo.split("-")
                    cod_color = parts[-1]; modelo = "-".join(parts[:-1])
                if color_split and re.match(r"^\d+-", color):
                    d = color.index("-")
                    cod_color = color[:d]; color = color[d+1:].strip()
            preview.append({
                "Marca": m, "Modelo": modelo, "Color": color,
                "Cód.Color": cod_color, "Cajas": item["cajas"],
                "Tallas": " ".join(map(str, item.get("tallas") or [])),
                "Cantidades": " ".join(map(str, item.get("dist") or [])),
                "Precio USD": item["precio"],
            })
        st.dataframe(preview, use_container_width=True, hide_index=True)

    # ── Generar Excel ────────────────────────────────────────────────────────
    st.divider()

    def _build_excel(items, marca, color_split, model_split, por_linea_marca):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Calzado"
        headers = ["ID Detalle","Marca","Modelo","Color","Cód. Color",
                   "Cajas","Tallas","Cantidades","Precio Unitario"]
        hfill = PatternFill("solid", fgColor="1A1A1A")
        hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        ctr   = Alignment(horizontal="center", vertical="center")
        thin  = Side(style="thin", color="E5E3DC")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hfill; cell.font = hfont
            cell.alignment = ctr; cell.border = brd

        alt     = PatternFill("solid", fgColor="F8F7F4")
        nrm     = Font(name="Arial", size=10)
        prc_fill= PatternFill("solid", fgColor="E1F5EE")

        for i, item in enumerate(items, 2):
            if "modelo_forzado" in item:
                modelo    = item["modelo_forzado"]
                cod_color = item.get("cod_color_forzado","")
                color     = item.get("color","")
                m         = item.get("marca_linea", marca)
            else:
                modelo, color, cod_color = item["ref"], item.get("color",""), ""
                m = marca
                if model_split and "-" in modelo:
                    parts = modelo.split("-")
                    cod_color = parts[-1]; modelo = "-".join(parts[:-1])
                if color_split and re.match(r"^\d+-", color):
                    d = color.index("-")
                    cod_color = color[:d]; color = color[d+1:].strip()

            tallas_str = " ".join(map(str, item.get("tallas") or []))
            dist_str   = " ".join(map(str, item.get("dist") or []))
            fill = alt if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            vals = ["", m, modelo, color, cod_color,
                    item["cajas"], tallas_str, dist_str, item["precio"]]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=val)
                cell.font = nrm; cell.border = brd
                cell.alignment = ctr if c in (1,2,5,6,7,8,9) else Alignment(vertical="center")
                cell.fill = prc_fill if c == 9 else fill
                if c == 9: cell.number_format = '"$"#,##0.00'

        for i, w in enumerate([10,12,16,38,12,8,18,16,14], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return buf

    if st.button("⬇  Generar y descargar Excel", type="primary", use_container_width=True):
        with st.spinner("Generando Excel..."):
            buf = _build_excel(items, marca, color_split, model_split, por_linea_marca)
        fname = f"{(meta.get('marca_archivo') or marca or 'importacion').lower()}_importacion.xlsx"
        st.download_button(
            label="📥 Descargar Excel",
            data=buf,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.balloons()

# ── Footer ────────────────────────────────────────────────────────────────────
st.divider()
st.caption("Generador Excel · Calzado — Soporta: PE COM PE · RAMARIM · ALA · GRENDENE · POLO GO · LEVETERAPIA · SANDRA · TREESHOES")
